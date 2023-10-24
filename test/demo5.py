# this file create basic strucure and add hyperlink

import openpyxl
from openpyxl.styles import Font

# 打开文件
file_path = "/Users/chang/Desktop/test123.xlsx"
workbook = openpyxl.load_workbook(file_path)

# 获取 "lot" 工作表
sheet_lot = workbook["lot"]

# 定义命名规则的各个部分的范围
letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
numbers1 = list(map(str, range(1, 13)))

# # 循环创建符合命名规则的工作表
for letter in letters:
    for number in numbers1:
        #         # 创建新工作表
        new_sheet_name = f"{letter}{number}"
        new_sheet = workbook.copy_worksheet(sheet_lot)
        new_sheet.title = new_sheet_name

# 为每个库位创建sheet
# 获取 "data" 工作表
sheet_data = workbook["data"]

# 定义命名规则的各个部分的范围
numbers2 = list(map(str, range(1, 5)))
lr = ["L", "R"]

# 循环创建符合命名规则的工作表
for letter in letters:
    for number1 in numbers1:
        for number2 in numbers2:
            for lr_value in lr:
                # 创建新工作表
                new_sheet_name = f"{letter}{number1}F{number2}{lr_value}"
                new_sheet = workbook.copy_worksheet(sheet_data)
                new_sheet.title = new_sheet_name

# 给lot添加超链接
# 获取第一个工作表（Sheet1）
sheet1 = workbook["Main Layer"]
# 创建样式对象，设置下划线和颜色
font_style = Font(underline="single", color="0000FF")

# 初始化目标单元格位置
target_row = None
target_column = None

# 循环为每个lot添加超链接到主页
for letter in letters:
    for number in numbers1:
        new_target_value = f"{letter}{number}"
        # 遍历前50行和前50列,查找目标lot
        for row in sheet1.iter_rows(max_row=50):
            for cell in row:
                if cell.value == new_target_value:
                    target_row = cell.row
                    target_column = cell.column
                    target_cell = sheet1.cell(row=target_row, column=target_column)
                    target_cell.font = font_style
                    target_cell.hyperlink = f"#{new_target_value}!A1"
                    break

# 给库位货架添加超链接
# 定义层数范围
letters2 = ["B", "C"]
numbers3 = list(map(str, range(2, 5)))

# 循环为每个lot添加超链接到主页
for letter in letters:
    for number in numbers1:
        # 找到要添加超链接的sheet
        sheet_lot = workbook[f"{letter}{number}"]
        # 定义要添加超链接的单元格范围
        cell_range = sheet_lot["B2:C5"]
        # 初始化层数为4
        layer_value = 4

        # 遍历位置添加超链接
        for row in cell_range:
            for cell in row:
                target_row = cell.row
                target_column = cell.column
                layer_value_str = str(layer_value)
                target_cell = sheet_lot.cell(row=target_row, column=target_column)
                target_cell.font = font_style
                leftOrRight = "L" if target_column == 2 else "R"
                if letter == "C":
                    target_cell.hyperlink = f"#'{letter}{number}F{layer_value_str}{leftOrRight}'!A1"
                else:
                    target_cell.hyperlink = f"#{letter}{number}F{layer_value_str}{leftOrRight}!A1"
            layer_value -= 1

# 保存工作簿
workbook.save(file_path)
